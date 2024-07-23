from PyQt5 import Qt, QtWidgets, QtGui
from PyQt5.QtCore import Qt as Qtt, QSize
from PyQt5.QtWidgets import QMessageBox, QInputDialog
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from app.db_requests import check_weight, get_id_by_name, get_nds_price_by_name, get_all_names


class AddDelScrap(Qt.QDialog):

    def __init__(self, flag, parent=None):
        super().__init__(parent)
        self.flag = flag
        self.result_dict = None
        self.row_count = None
        self.name_list = get_all_names()
        self.name_list.insert(0, "Выберите наименование")
        self.setGeometry(560, 240, 800, 600)
        if flag:
            self.setWindowTitle('Прибытие')
        else:
            self.setWindowTitle('Убытие')
        self.setWindowIcon(QtGui.QIcon("icons/Icon.png"))

        self.group_list = []
        if flag:
            self.label = Qt.QLabel('Выберите прибывшие наименование(-я)')
        else:
            self.label = Qt.QLabel('Выберите убывшие наименование(-я)')
        self.label.setStyleSheet("color:black; font: bold 20pt 'Arial';")
        self.label.setAlignment(Qtt.AlignCenter)

        self.table = Qt.QTableWidget()

        self.start = Qt.QPushButton('Сохранить и выйти')

        self.v_layout = Qt.QVBoxLayout(self)
        self.v_layout.addWidget(self.label)
        self.v_layout.addWidget(self.table)
        self.v_layout.addWidget(self.start)

        self.name_table()

        self.start.clicked.connect(self.go_out_btn)

    def name_table(self):
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Наименование", "Количество", "Удалить"])

        self.row_count = self.table.rowCount()
        self.table.insertRow(self.row_count)

        self.row_items()

        self.table.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
        self.table.horizontalHeader().setDefaultAlignment(Qtt.AlignCenter)
        self.table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)

    def row_items(self):
        item = Qt.QComboBox()
        item.addItems(self.name_list)
        item.currentTextChanged.connect(self.add_new_row)
        self.table.setCellWidget(self.row_count, 0, item)

        item = Qt.QLineEdit()
        item.setDisabled(True)
        item.setAlignment(Qtt.AlignCenter)
        self.table.setCellWidget(self.row_count, 1, item)

        item = Qt.QPushButton()
        item.setIcon(QtGui.QIcon("icons/Del.png"))
        item.setIconSize(QSize(22, 22))
        item.clicked.connect(self.del_btn)
        item.setDisabled(True)
        self.table.setCellWidget(self.row_count, 2, item)

    def add_new_row(self):
        self.table.cellWidget(self.row_count, 0).setDisabled(True)
        self.table.cellWidget(self.row_count, 1).setDisabled(False)
        self.table.cellWidget(self.row_count, 2).setDisabled(False)
        name = self.table.cellWidget(self.row_count, 0).currentText()
        self.name_list.remove(name)
        self.row_count = self.table.rowCount()
        self.table.insertRow(self.row_count)
        self.row_items()

    def go_out_btn(self):
        self.result_dict = {}
        try:
            for x in range(self.row_count):
                key = self.table.cellWidget(x, 0).currentText()
                value = float(self.table.cellWidget(x, 1).text())
                if not value:
                    raise ValueError
                elif not self.flag and not check_weight(key, value):
                    raise Warning
                else:
                    self.result_dict[key] = value
        except ValueError as e:
            print(e)
            Qt.QMessageBox.critical(self, 'Ошибка!', 'В поле количество введите массу в тоннах!\nДробная часть '
                                                     'вводится через точку.')
        except Warning as e:
            print(e)
            Qt.QMessageBox.critical(self, 'Ошибка!', f'Вес наименования {key} превышает доступный в наличии!')
        else:
            if self.result_dict:
                button = QMessageBox.question(self, "Создание отчета", "Сохранить отчет об изменениях?")
                if button == QMessageBox.Yes:
                    name = ''
                    while not name:
                        name, ok = QInputDialog.getText(self, 'Новый отчет', 'Введите название файла:')
                        if ok:
                            if name:
                                self.edit_report(name)
                            else:
                                Qt.QMessageBox.critical(self, 'Ошибка!', 'Название файла не может быть пустым!')
                        else:
                            break

                self.accept()
            else:
                self.reject()

    def del_btn(self):
        name = self.table.cellWidget(self.table.currentRow(), 0).currentText()
        self.name_list.append(name)
        self.table.update()
        self.table.removeRow(self.table.currentRow())
        self.row_count = self.table.rowCount() - 1

        self.upd_list()

    def upd_list(self):
        self.name_list.remove("Выберите наименование")
        self.name_list.sort()
        self.name_list.insert(0, "Выберите наименование")

        item = Qt.QComboBox()
        item.addItems(self.name_list)
        item.currentTextChanged.connect(self.add_new_row)
        self.table.setCellWidget(self.row_count, 0, item)

    def edit_report(self, name):
        wb = Workbook()

        ws = wb.active
        if self.flag:
            ws.title = "Прибытие"
        else:
            ws.title = "Убытие"

        titles = ["ID", "Наименование", "Количество, т.", "Цена, руб.", "Сумма, руб.", "% НДС", "НДС, руб.",
                  "Всего, руб."]
        widths = [3, 40, 20, 20, 20, 11, 11, 20]

        for x in range(1, 9):
            current_cell = ws.cell(row=1, column=x, value=titles[x - 1])
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            letter = get_column_letter(x)
            ws.column_dimensions[letter].width = widths[x - 1]

        y = 2
        for one in self.result_dict.keys():
            current_cell = ws.cell(row=y, column=1, value=get_id_by_name(one))
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell = ws.cell(row=y, column=2, value=one)
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell = ws.cell(row=y, column=3, value=self.result_dict[one])
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            nds_n_price = get_nds_price_by_name(one)
            current_cell = ws.cell(row=y, column=4, value=nds_n_price.price)
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            cost = float(format(nds_n_price.price * self.result_dict[one], '.2f'))
            current_cell = ws.cell(row=y, column=5, value=cost)
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell = ws.cell(row=y, column=6, value=nds_n_price.percent_nds)
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            nds = float(format(nds_n_price.percent_nds * cost * 0.01, '.2f'))
            current_cell = ws.cell(row=y, column=7, value=nds)
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell = ws.cell(row=y, column=8, value=cost - nds)
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            y += 1

        wb.save(f'{name}.xlsx')
