from openpyxl import Workbook

from app.db_requests import all_table, scrap_names


def create_report(name):

    wb = Workbook()

    ws = wb.active

    ws.title = "Report"

    records = all_table()

    titles = ["ID", "Наименование", "Количество", "Цена", "Сумма", "% НДС", "НДС", "Всего"]

    for x in range(1, 9):
        ws.cell(row=1, column=x, value=titles[x-1])

    y = 2
    for row in records:
        ws.cell(row=y, column=1, value=row.id)
        ws.cell(row=y, column=2, value=scrap_names(row.id))
        ws.cell(row=y, column=3, value=row.weight)
        ws.cell(row=y, column=4, value=row.price)
        cost = float(format(row.price * row.weight, '.2f'))
        ws.cell(row=y, column=5, value=cost)
        ws.cell(row=y, column=6, value=row.percent_nds)
        nds = float(format(row.percent_nds * cost * 0.01, '.2f'))
        ws.cell(row=y, column=7, value=nds)
        ws.cell(row=y, column=8, value=cost - nds)
        y += 1

    wb.save(f'{name}.xlsx')
