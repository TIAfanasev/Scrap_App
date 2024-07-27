from PyQt5 import Qt, QtWidgets, QtGui
from PyQt5.QtCore import Qt as Qtt
from PyQt5.QtWidgets import QTableWidgetItem, QHeaderView
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment

from db_requests import all_table


class Report(Qt.QDialog):

    def __init__(self, filename, parent=None):
        super().__init__(parent)

        self.filename = filename
        self.setGeometry(560, 240, 300, 600)
        self.setWindowTitle('Отчет')
        self.setWindowIcon(QtGui.QIcon("icons/Icon.png"))
        self.label = Qt.QLabel('Список позиций')
        self.label.setStyleSheet("color:black; font: bold 20pt 'Arial';")
        self.label.setAlignment(Qtt.AlignCenter)

        self.table = Qt.QTableWidget()

        self.confirm_btn = Qt.QPushButton('Создать отчет')

        self.v_layout = Qt.QVBoxLayout(self)
        self.v_layout.addWidget(self.label)
        self.v_layout.addWidget(self.table)
        self.v_layout.addWidget(self.confirm_btn)

        self.confirm_btn.clicked.connect(self.create_btn_click)

        self.table_filling()

    def table_filling(self):
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(
            ["Наименование", "Выбрать"])

        records = all_table()

        for row in records:
            row_count = self.table.rowCount()
            self.table.insertRow(row_count)

            item = QTableWidgetItem(str(row.ScrapNameList.name))
            item.setTextAlignment(Qtt.AlignCenter)
            self.table.setItem(row_count, 0, item)

            item = Qt.QCheckBox()
            self.table.setCellWidget(row_count, 1, item)

            self.table.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)

            self.table.horizontalHeader().setDefaultAlignment(Qtt.AlignCenter)
            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
            self.table.resizeRowsToContents()

    def create_btn_click(self):
        row_count = self.table.rowCount()
        name_list = []
        for x in range(row_count):
            if self.table.cellWidget(x, 1).isChecked():
                name_list.append(self.table.item(x, 0).text())
        self.create_report(name_list)
        self.accept()

    def create_report(self, name_list):
        wb = Workbook()

        ws = wb.active

        ws.title = "Report"

        records = all_table()

        titles = ["ID", "Наименование", "Количество, т.", "Цена, руб.", "Сумма, руб.", "% НДС", "НДС, руб.", "Всего, руб."]
        widths = [3, 40, 20, 20, 20, 11, 11, 20]

        for x in range(1, 9):
            current_cell = ws.cell(row=1, column=x, value=titles[x - 1])
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            letter = get_column_letter(x)
            ws.column_dimensions[letter].width = widths[x-1]

        y = 2
        for row in records:
            if row.ScrapNameList.name in name_list:
                current_cell = ws.cell(row=y, column=1, value=row.ScrapList.id)
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_cell = ws.cell(row=y, column=2, value=row.ScrapNameList.name)
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_cell = ws.cell(row=y, column=3, value=row.ScrapList.weight)
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_cell = ws.cell(row=y, column=4, value=row.ScrapList.price)
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                cost = float(format(row.ScrapList.price * row.ScrapList.weight, '.2f'))
                current_cell = ws.cell(row=y, column=5, value=cost)
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_cell = ws.cell(row=y, column=6, value=row.ScrapList.percent_nds)
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                nds = float(format(row.ScrapList.percent_nds * cost * 0.01, '.2f'))
                current_cell = ws.cell(row=y, column=7, value=nds)
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_cell = ws.cell(row=y, column=8, value=cost - nds)
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                y += 1

        wb.save(f'reports\\{self.filename}.xlsx')
